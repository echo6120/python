#coding=utf-8

import requests
import urllib3
import re
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

#爬取课程名称，价格，开课时间，主讲老师
def get_courseinfo(kelink_file):
    result=[]
    with open(kelink_file) as fp:
            for keurl in fp:
                urllib3.disable_warnings()
                resq2 = requests.get(keurl.strip(),verify=False).text
                soup = BeautifulSoup(resq2,'html.parser')
                try:
                    names = soup.select("div.g-w.body > div > h1")
                    ##mod-course-head > div.g-w.body > div > h1
                    teachernames = soup.select("div.g-w.body > div > p")
                    coursetimes = soup.select("div.g-w.body > div > p")
                    if names!=None and teachernames!=None and coursetimes!=None:
                        data = {
                            'name': str(names[0]).strip().strip('<h1>').strip('</h1>'),
                            'teachername:': str(teachernames[0]).strip().strip("<p>").strip("</p>"),
                            'coursetimes': str(coursetimes[1]).strip().strip("<p>").strip("</p>"),
                            'url': keurl.strip()
                        }
                        result.append(data)
                        #print data
                    else:
                        print u"有属性为空了,skip skip"
                except Exception, e:
                        print e
            return result

#将爬虫下来的内容保存在Excel
def write_excel(filename,result):
    wb = load_workbook(filename)
    wb.guess_types = True
    ws=wb.active
    #excel表中有多少行，Excel的行和列是从第一行列开始的
    for i in range(1,len(result)+1):
        #取result列表中的每个data，每个data为一行
        result_item = result[i-1]
        #默认从第一列开始
        column_num=1
        #遍历字典data，每个属性增加一列
        for key,value1 in result_item.items():
            if value1.strip() != None:
                ws.cell(row=i,column=column_num,value=value1)
                column_num+=1
    wb.save(filename)

kelink_file = "d:\\kelink.txt"
#抓取课程详细信息
get_courseinfo(kelink_file)
result = get_courseinfo(kelink_file)[1:6]
write_excel("d:\\test.xlsx",result)